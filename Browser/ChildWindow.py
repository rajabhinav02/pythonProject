import os
import time
import openpyxl
from selenium import webdriver

driver= webdriver.Chrome(executable_path="C:\\chromedriver_win32\\chromedriver.exe")

driver.get("https://the-internet.herokuapp.com/windows")

#driver.find_element_by_css_selector("a[href*='windows']").click()

driver.find_element_by_xpath("//a[contains(@href,'windows')]").click()
driver.switch_to.window(driver.window_handles[1])

print(driver.find_element_by_css_selector("h3").text)
driver.close()

driver.switch_to.window(driver.window_handles[0])
print(driver.find_element_by_xpath("//h3").text)

assert "Opening" in driver.find_element_by_xpath("//h3").text

driver.execute_script("window.scroll(0,-1000);")
#driver.execute_script("arguments[0].scrollIntoView(true);", element)

def screenshot(testmessage):
    ssfilename = testmessage+ str(time.time()*1000) +".png"
    ssdirectory = "../Screenshot/"
    relfilename = ssdirectory+ssfilename
    ccdir = os.path.dirname(__file__)
    acfilename =os.path.join(ccdir,relfilename)
    acdirectory = os.path.join(ccdir,ssdirectory)

    try:
        if not os.path.exists(acdirectory):
            os.makedirs(acdirectory)
            driver.get_screenshot_as_file(acfilename)
    except:
        print("lol")

class TetStatus:
    testlist = []
    def teststatus(self, result, testmessage):
        if result is not None:
            if result:
                self.testlist.append("Pass")
            else:
                self.testlist.append("Fail")
        else:
            self.testlist.append("Fail")

    def markstatus(self, result, testmessage):
        self.teststatus(result, testmessage)

    def marktestfinal(self, result, testmessage, tcname):
        self.teststatus(result, testmessage)

        if "Fail" in self.testlist:
            self.testlist.clear()
            assert True== False
        else:
            self.testlist.clear()
            assert True==True

class excel:
    def openexcel(self, tcname):
        dict= {}
        book = openpyxl.load_workbook("sheetname")
        sheet= book.active

        for i in (1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == tcname:
                for j in (2, sheet.max_column+1):
                    dict[sheet.cell(row=1, column =j).value] = sheet.cell(row=i, column = j).value

        return [dict]


